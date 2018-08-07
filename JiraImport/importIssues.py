from jira import JIRA
#from xlsxwriter.workbook import Workbook
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter

date = time.strftime('%Y%m%d%H%M%S')

# <!----- PARAMETERS ------
# These parameters are imported from myconfig.py
'''
JIRA_BASE_URL = 'https://jira.org.co.nz'
ConfigFile ="config.dat"
SprintExtract = "Sprints"
JiraExtract = "JiraIssues"
WorkLogExtract = "WorkLogs"
ReleasesExtract = "Releases"
TeamMemberExtract = 'TeamMembers'

date = time.strftime('%Y%m%d%H%M%S')
xlext = '.xlsx'
csvext = '.csv'

# Jira Issue field list
FieldList = ['issuetype', 'project', 'status', 'resolution', 'created', 'timeestimate',
                 'aggregatetimeoriginalestimate', 'aggregatetimeestimate',
                 'timespent', 'aggregatetimespent', 'resolutiondate', 'customfield_10000', 'customfield_10001',
                 'customfield_11412', 'customfield_10103', 'customfield_10600','fixVersions', 'customfield_10008', 'summary', 'priority', 'customfield_10400']

# Sprint field list
SPFieldList = ['rapidViewId', 'state', 'name', 'startDate', 'endDate', 'completeDate', 'sequence']

# Sprints: Fields to removed
spfieldremove= ['rapidViewId=', 'state=', 'name=', 'startDate=', 'endDate=', 'completeDate=', 'sequence=']

# Work Log field list
WLFieldList = ['issuekey','id', 'issueId', 'created','author.name', 'timeSpentSeconds']
#,'runningremainingestimate','totalremaining', 'cummtimespent'

# Members fields list
MembersFieldList =['id', 'name', 'key', 'displayname', 'availability', 'team', 'teamname']

# Releases Fields List
ReleasesFieldList = ['id', 'name', 'released', 'releaseDate', 'projectId']
'''
# ----- PARAMETERS ------>

def List_all_Fields():
    jira = SessionSetup(1)
    issue = jira.issue('SWAG2-10177')
    for field_name in issue.raw['fields']:
        # print("Field:", field_name, "Value:", issue.raw['fields'][field_name])
        print("Field:{0}, Value:{1}".format(field_name, issue.raw['fields'][field_name]))


def listallboards():
    jira=SessionSetup(1)
    issue = jira.issue('SWAG2-2522')
    jt=jira.transitions(issue)
    p=jira.project(issue.fields.project)

    boards = jira.boards()
    for board in boards:
        print('{0} : {1}'.format(str(board.id).ljust(5), board.name))


    #for f in p.raw['fields']:
    #    print(p.raw['fields'][f])


def getConfig(confvar):
    projectList=''

    if (not os.path.exists(ConfigFile)):
        print("Whhoops! config file not found ", ConfigFile)
    else:

        f = open(ConfigFile, 'r+')
        file_data = f.read().splitlines()
        f.close()

        for line in file_data:
            if line.startswith(confvar):
                projectList = line.replace(confvar, '').replace('=', '').split(',')
                break
    return projectList


def main(argv):

    userName = ''
    password = ''
    basefilename = os.path.basename(__file__)

    try:
        opts, args = getopt.getopt(argv, "hu:p:", ["uname=", "pass="])
    except getopt.GetoptError:
        print (basefilename, ' -u <username> -p <password>')
        sys.exit(2)
    for opt, arg in opts:
        if opt == '-h':
            print (basefilename, ' -u <username> -p <password>')
            sys.exit()
        elif opt in ("-u", "--uname"):
            userName = arg
        elif opt in ("-p", "--pass"):
            password = arg

    if not userName or not password :
        print(basefilename, ' -u <username> -p <password>')
        sys.exit()
    else:
        global UNAME
        global PASSWD

        UNAME=userName
        PASSWD=password

        executeExtractProcess()
    '''
    global UNAME
    global PASSWD
    UNAME = 'xxxx'
    PASSWD = 'xxxx'
    executeExtractProcess()
    '''

if __name__ == '__main__':
    urllib3.disable_warnings()
    main(sys.argv[1:])

